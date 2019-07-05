## Creates PLO to CLO mapping files CAC Programs
# Peter Ryan 28 Feb 201

# v1.6
# Uses audited PLOs and CLOs
from typing import Union

import pandas as pd
from openpyxl.worksheet import Worksheet
from openpyxl.worksheet.write_only import WriteOnlyWorksheet
from tabulate import tabulate

import openpyxl
from openpyxl.styles import Protection
from openpyxl.styles import Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Color, PatternFill, Font, Border

import sys
sys.path.append('c:\\Peter\\GitHub\\CoB\\')

import general.RMIT_colours as rc

from datetime import datetime
startTime = datetime.now()
print(datetime.now().strftime('%H:%M:%S %d/%b/%y'))

# Fills
cloFill = PatternFill(fgColor=rc.RMIT_Blue3[1:], fill_type='solid')
ploFill = PatternFill(fgColor=rc.RMIT_Yellow3[1:], fill_type='solid')
dateFill = PatternFill(fgColor=rc.RMIT_Orange3[1:], fill_type='solid')


def adjust_excel_rows(ws, skip_rows=0, max_rows=False, default_col='A', default_height=15):
  width = ws.column_dimensions[default_col].width
  for i_row, row in enumerate(ws):
    if (i_row + 1 > skip_rows and
      (max_rows == False or max_rows >= (i_row + 1 - skip_rows))):
      multiples_of_font_size = [1]
      for j_col, cell in enumerate(row):
        if cell.value is not None:
          mul = 0
          for v in str(cell.value).split('\n'):
            mul += int(len(v)/width) + 1
          if mul > 0:
            multiples_of_font_size.append(mul)
          
        cell.alignment = Alignment(wrap_text=True)
      
      new_height = max(multiples_of_font_size) * default_height
      ws.row_dimensions[i_row + 1].height = new_height


# open template
directory = 'H:\\Projects\\CoB\\Program Transformation\\CLO mapping\\Success\\'

plo_clo_filename = 'Audit\\PLO_CLO_alignment_all_programs_28Feb2019_complete_audit.xlsx'

template = 'PLO_CLO_alignment_Feb28_temp.xlsx'

# Get Dataframes
df_plo = pd.read_excel(directory+plo_clo_filename,
                       sheet_name='PLOs_audit',
                       converters={'General': str,
                                   'General.1': str,
                                   'Active Verb': str,
                                   'Active Verb.1': str,
                                   'Focus/Object': str,
                                   'Focus/Object.1': str,
                                   'Context/Qualifier': str,
                                   'Context/Qualifier.1': str,
                                   'Clear and usable': str,
                                   'Comments': str
                                   })
df_clo = pd.read_excel(directory+plo_clo_filename,
                       sheet_name='CLOs_audit',
                       converters={'course_id': str,
                                   'General': str,
                                   'General.1': str,
                                   'Active Verb': str,
                                   'Active Verb.1': str,
                                   'Focus/Object': str,
                                   'Focus/Object.1': str,
                                   'Context/Qualifier': str,
                                   'Context/Qualifier.1': str,
                                   'Clear and usable': str,
                                   'Comments': str
                                   })
df_mapping = pd.read_excel(directory+plo_clo_filename, sheet_name='Program_course_mapping', converters={'course_id': str})


#print(tabulate(df_plo[:20], headers='keys'))
#print(tabulate(df_clo[:20], headers='keys'))
#print(tabulate(df_mapping[:20], headers='keys'))

# get program list
df_programs = df_mapping[['program_code', 'plan_code', 'program_name']].drop_duplicates()

# Matching column data validation
dv_match = DataValidation(type="list", formula1='"0,1"', allow_blank=True)

# Optionally set a custom error message
dv_match.error ='Your entry is not in the list'
dv_match.errorTitle = 'Invalid Entry'
# Optionally set a custom prompt message
dv_match.prompt = 'Please select from the list'
dv_match.promptTitle = 'List Selection'


# General column data validation
dv_general = DataValidation(type="list", formula1='selections!$A$2:$A$7', allow_blank=True)

# Optionally set a custom error message
dv_general.error ='You have entered free text'
dv_general.errorTitle = 'Free Text'
dv_general.errorStyle = 'warning'

# Active Verb column data validation
dv_active = DataValidation(type="list", formula1='selections!$B$2:$B$6', allow_blank=True)

# Optionally set a custom error message
dv_active.error ='You have entered free text'
dv_active.errorTitle = 'Free Text'
dv_active.errorStyle = 'warning'

# Focus column data validation
dv_focus = DataValidation(type="list", formula1='selections!$C$2:$C$6', allow_blank=True)

# Optionally set a custom error message
dv_focus.error ='You have entered free text'
dv_focus.errorTitle = 'Free Text'
dv_focus.errorStyle = 'warning'

# Context column data validation
dv_context = DataValidation(type="list", formula1='selections!$D$2:$D$7', allow_blank=True)

# Optionally set a custom error message
dv_context.error ='You have entered free text'
dv_context.errorTitle = 'Free Text'
dv_context.errorStyle = 'warning'

# Clear column data validation
dv_clear = DataValidation(type="list", formula1='selections!$E$2:$E$5', allow_blank=True)

# Optionally set a custom error message
dv_clear.error ='You have entered free text'
dv_clear.errorTitle = 'Free Text'
dv_clear.errorStyle = 'warning'

# General column data validation
dv_general2 = DataValidation(type="list", formula1='selections!$A$2:$A$7', allow_blank=True)

# Optionally set a custom error message
dv_general2.error ='You have entered free text'
dv_general2.errorTitle = 'Free Text'
dv_general2.errorStyle = 'warning'

# Active Verb column data validation
dv_active2 = DataValidation(type="list", formula1='selections!$B$2:$B$6', allow_blank=True)

# Optionally set a custom error message
dv_active2.error ='You have entered free text'
dv_active2.errorTitle = 'Free Text'
dv_active2.errorStyle = 'warning'

# Focus column data validation
dv_focus2 = DataValidation(type="list", formula1='selections!$C$2:$C$6', allow_blank=True)

# Optionally set a custom error message
dv_focus2.error ='You have entered free text'
dv_focus2.errorTitle = 'Free Text'
dv_focus2.errorStyle = 'warning'

# Context column data validation
dv_context2 = DataValidation(type="list", formula1='selections!$D$2:$D$7', allow_blank=True)

# Optionally set a custom error message
dv_context2.error ='You have entered free text'
dv_context2.errorTitle = 'Free Text'
dv_context2.errorStyle = 'warning'

# Clear column data validation
dv_clear2 = DataValidation(type="list", formula1='selections!$E$2:$E$5', allow_blank=True)

# Optionally set a custom error message
dv_clear2.error ='You have entered free text'
dv_clear2.errorTitle = 'Free Text'
dv_clear2.errorStyle = 'warning'

for i_prg, prg in df_programs[40:].iterrows():
  # initiate counters
  j_plo = 2
  j_mapping = 2
  j_clo = 2
  j_align = 2
  print(prg['program_code'])
  # open template
  wb = openpyxl.load_workbook(directory + template)

  # Open sheets
  ws_plo = wb["PLOs"]
  ws_mapping = wb["Program_course_mapping"]
  ws_clo = wb["CLOs"]
  ws_align = wb["PLOs_to_CLOs"]
  ws_all_courses = wb['All Course Summaries']
  
  # Add the data-validation object to the align worksheet
  ws_align.add_data_validation(dv_match)
  
  # Add data-validation to plo worksheet
  ws_plo.add_data_validation(dv_general)
  ws_plo.add_data_validation(dv_active)
  ws_plo.add_data_validation(dv_focus)
  ws_plo.add_data_validation(dv_context)
  ws_plo.add_data_validation(dv_clear)

  # Add data-validation to clo worksheet
  ws_clo.add_data_validation(dv_general2)
  ws_clo.add_data_validation(dv_active2)
  ws_clo.add_data_validation(dv_focus2)
  ws_clo.add_data_validation(dv_context2)
  ws_clo.add_data_validation(dv_clear2)
  
  # Filter dataframes to program level
  df_prg_plos = df_plo.loc[(df_plo['program_code'] == prg['program_code']) & (df_plo['plan_code'] == prg['plan_code'])]
  df_prg_mapping = df_mapping.loc[
    (df_mapping['program_code'] == prg['program_code']) & (df_mapping['plan_code'] == prg['plan_code'])]
  df_prg_courses = df_prg_mapping[['course_id', 'course_code', 'course_name']].drop_duplicates()
  
  
  # Update PLO sheet & All course summaries
  for i, plo in df_prg_plos.iterrows():
    ws_plo.cell(row=j_plo, column=1).value = prg['program_code']
    ws_plo.cell(row=j_plo, column=2).value = prg['plan_code']
    ws_plo.cell(row=j_plo, column=3).value = prg['program_name']
    ws_plo.cell(row=j_plo, column=4).value = plo['career']
    ws_plo.cell(row=j_plo, column=5).value = plo['school_code']
    ws_plo.cell(row=j_plo, column=6).value = plo['school_abbr']
    ws_plo.cell(row=j_plo, column=7).value = plo['plo_nbr']
    ws_plo.cell(row=j_plo, column=8).alignment = Alignment(wrapText=True)
    ws_plo.cell(row=j_plo, column=8).value = plo['plo_text']
    
    if plo['General'] != 'NaN':
      ws_plo.cell(row=j_plo, column=9).value = plo['General']
    ws_plo.cell(row=j_plo, column=10).value = plo['General.1']
    ws_plo.cell(row=j_plo, column=11).value = plo['Active Verb']
    ws_plo.cell(row=j_plo, column=12).value = plo['Active Verb.1']
    ws_plo.cell(row=j_plo, column=13).value = plo['Focus/Object']
    ws_plo.cell(row=j_plo, column=14).value = plo['Focus/Object.1']
    ws_plo.cell(row=j_plo, column=15).value = plo['Context/Qualifier']
    ws_plo.cell(row=j_plo, column=16).value = plo['Context/Qualifier.1']
    ws_plo.cell(row=j_plo, column=17).value = plo['Clear and usable']
    ws_plo.cell(row=j_plo, column=18).value = plo['Comments']
    
    dv_general.add(ws_plo.cell(row=j_plo, column=9))
    dv_general.add(ws_plo.cell(row=j_plo, column=10))
    dv_active.add(ws_plo.cell(row=j_plo, column=11))
    dv_active.add(ws_plo.cell(row=j_plo, column=12))
    dv_focus.add(ws_plo.cell(row=j_plo, column=13))
    dv_focus.add(ws_plo.cell(row=j_plo, column=14))
    dv_context.add(ws_plo.cell(row=j_plo, column=15))
    dv_context.add(ws_plo.cell(row=j_plo, column=16))
    dv_clear.add(ws_plo.cell(row=j_plo, column=17))
    #ws_plo.cell(row=j_plo, column=9).value = r['status']
    #ws_plo.cell(row=j_plo, column=10).value = r['updated']

    ws_all_courses.cell(row=1, column=j_plo+3).alignment = Alignment(wrapText=True)
    ws_all_courses.cell(row=1, column=j_plo+3).value = plo['plo_nbr']
    ws_all_courses.cell(row=1, column=j_plo+3).fill = ploFill
    ws_all_courses.cell(row=2, column=j_plo+3).alignment = Alignment(wrapText=True)
    ws_all_courses.cell(row=2, column=j_plo+3).value = plo['plo_text']
    ws_all_courses.cell(row=2, column=j_plo+3).fill = ploFill
    j_plo += 1

  print('PLO Done: {}'.format(j_plo))
  
  # Update mapping sheet
  for i, r in df_prg_mapping.iterrows():
    df_prg_crse_clos = df_clo.loc[(df_clo['course_id'] == r['course_id'])]
    ws_mapping.cell(row=j_mapping, column=1).value = prg['program_code']
    ws_mapping.cell(row=j_mapping, column=2).value = prg['plan_code']
    ws_mapping.cell(row=j_mapping, column=3).value = prg['program_name']
    ws_mapping.cell(row=j_mapping, column=4).value = r['course_id']
    ws_mapping.cell(row=j_mapping, column=5).value = r['course_code']
    ws_mapping.cell(row=j_mapping, column=6).value = r['course_name']
    ws_mapping.cell(row=j_mapping, column=7).value = r['course_list_name']
    ws_mapping.cell(row=j_mapping, column=8).value = len(df_prg_crse_clos)
    j_mapping += 1
  print('Mapping Done: {}'.format(j_mapping))
  
  # Update CLOs sheet & All course summaries
  for i_crse, crse in df_prg_courses.iterrows():
    df_prg_crse_clos = df_clo.loc[(df_clo['course_id'] == crse['course_id'])]
    for i, clo in df_prg_crse_clos.iterrows():
      ws_clo.cell(row=j_clo, column=1).value = crse['course_id']
      ws_clo.cell(row=j_clo, column=2).value = crse['course_code']
      ws_clo.cell(row=j_clo, column=3).value = crse['course_name']
      ws_clo.cell(row=j_clo, column=4).value = clo['school_code']
      ws_clo.cell(row=j_clo, column=5).value = clo['school_abbr']
      ws_clo.cell(row=j_clo, column=6).value = clo['clo_nbr']
      ws_clo.cell(row=j_clo, column=7).alignment = Alignment(wrapText=True)
      ws_clo.cell(row=j_clo, column=7).value = clo['clo_text']

      if clo['General'] != 'NaN':
        ws_clo.cell(row=j_clo, column=8).value = clo['General']
      if clo['General.1'] != 'NaN':
        ws_clo.cell(row=j_clo, column=9).value = clo['General.1']
      if clo['Active Verb'] != 'NaN':
        ws_clo.cell(row=j_clo, column=10).value = clo['Active Verb']
      if clo['Active Verb.1'] != 'NaN':
        ws_clo.cell(row=j_clo, column=11).value = clo['Active Verb.1']
      if clo['Focus/Object'] != 'NaN':
        ws_clo.cell(row=j_clo, column=12).value = clo['Focus/Object']
      if clo['Focus/Object.1'] != 'NaN':
        ws_clo.cell(row=j_clo, column=13).value = clo['Focus/Object.1']
      if clo['Context/Qualifier'] != 'NaN':
        ws_clo.cell(row=j_clo, column=14).value = clo['Context/Qualifier']
      if clo['Context/Qualifier.1'] != 'NaN':
        ws_clo.cell(row=j_clo, column=15).value = clo['Context/Qualifier.1']
      if clo['Clear and usable'] != 'NaN':
        ws_clo.cell(row=j_clo, column=16).value = clo['Clear and usable']
      
      dv_general2.add(ws_clo.cell(row=j_clo, column=8))
      dv_general2.add(ws_clo.cell(row=j_clo, column=9))
      dv_active2.add(ws_clo.cell(row=j_clo, column=10))
      dv_active2.add(ws_clo.cell(row=j_clo, column=11))
      dv_focus2.add(ws_clo.cell(row=j_clo, column=12))
      dv_focus2.add(ws_clo.cell(row=j_clo, column=13))
      dv_context2.add(ws_clo.cell(row=j_clo, column=14))
      dv_context2.add(ws_clo.cell(row=j_clo, column=15))
      dv_clear2.add(ws_clo.cell(row=j_clo, column=16))
      #ws_clo.cell(row=j_clo, column=8).value = clo['version']
      #ws_clo.cell(row=j_clo, column=9).value = clo['status']
      #ws_clo.cell(row=j_clo, column=10).value = clo['published']
      
      # Update all course summary as well
      ws_all_courses.cell(row=j_clo+1, column=1).value = crse['course_code']
      ws_all_courses.cell(row=j_clo+1, column=1).fill = cloFill
      ws_all_courses.cell(row=j_clo+1, column=2).value = clo['clo_nbr']
      ws_all_courses.cell(row=j_clo+1, column=2).fill = cloFill
      ws_all_courses.cell(row=j_clo+1, column=3).alignment = Alignment(wrapText=True)
      ws_all_courses.cell(row=j_clo+1, column=3).value = clo['clo_text']
      ws_all_courses.cell(row=j_clo+1, column=3).fill = cloFill
      #ws_all_courses.cell(row=j_clo+1, column=4).value = clo['published']
      #ws_all_courses.cell(row=j_clo+1, column=4).fill = dateFill
      j_clo += 1
  print('CLOs Done: {}'.format(j_clo))
  
  # Update CLO & PLOS to CLOs sheets
  # For each plo iterate through all clos

  for i_plo, plo in df_prg_plos.iterrows():
    for i_crse, crse in df_prg_courses.iterrows():
      df_prg_crse_clos = df_clo.loc[(df_clo['course_id'] == crse['course_id'])]
      for i, clo in df_prg_crse_clos.iterrows():
        
        ws_align.cell(row=j_align, column=1).value = plo['plo_nbr']
        ws_align.cell(row=j_align, column=2).value = clo['clo_nbr']
        ws_align.cell(row=j_align, column=3).value = plo['program_code']
        ws_align.cell(row=j_align, column=4).value = crse['course_code']
        ws_align.cell(row=j_align, column=5).alignment = Alignment(wrapText=True)
        ws_align.cell(row=j_align, column=5).value = plo['plo_text']
        ws_align.cell(row=j_align, column=6).alignment = Alignment(wrapText=True)
        ws_align.cell(row=j_align, column=6).value = clo['clo_text']
        ws_align.cell(row=j_align, column=7).value = plo['Clear and usable']
        ws_align.cell(row=j_align, column=8).value = clo['Clear and usable']
        
        ws_align.cell(row=j_align, column=9).protection = Protection(locked=False)
        dv_match.add(ws_align.cell(row=j_align, column=9))
        j_align += 1

  print('Align Done: {}'.format(j_align))
  
  intTime1 = datetime.now()
  print('Time Taken:\t{}'.format(intTime1 - startTime))
  intTime2 = intTime1
  
  # Adjust row heights and column widths
  ws = wb["Program Summary"]
  ws.column_dimensions['B'].width = 60
  for i_row, row in enumerate(ws):
    if i_row + 1 > 3 and not i_row + 1 >= 16:
      # adjust height of row
      ws.row_dimensions[i_row + 1].height = 65
      for j_col, cell in enumerate(row):
        # centre text
        cell.alignment = Alignment(horizontal='center', vertical='center')
        # Wrap column B
        if j_col + 1 == 2:
          cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
  
  ws = wb["PLOs"]
  ws.column_dimensions['H'].width = 80
  adjust_excel_rows(ws, skip_rows=1, default_col='H', default_height=15)
  
  ws = wb["CLOs"]
  ws.column_dimensions['G'].width = 80
  adjust_excel_rows(ws, skip_rows=1, default_col='G', default_height=15)
  
  ws = wb["PLOs_to_CLOs"]
  ws.column_dimensions['E'].width = 80
  ws.column_dimensions['F'].width = 80
  adjust_excel_rows(ws, skip_rows=1, default_col='E', default_height=16)
  
  # Alter column and row dimensions in All courses
  ws_all_courses.row_dimensions[2].height = 200
  for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
    ws_all_courses.column_dimensions[col].width = 16
  
  adjust_excel_rows(ws_all_courses, skip_rows=2, default_col='C', default_height=15)
  print('Adjustments Done')
  
  intTime1 = datetime.now()
  print('Time Taken:\t{}'.format(intTime1 - intTime2))
  intTime2 = intTime1
  
  # Delete unnecessary rows from All courses
  row_no = 3
  while ws_all_courses.cell(row=row_no, column=1).value != None:
    row_no += 1
    
  for count in range(305-row_no):
    ws_all_courses.delete_rows(row_no)

  print('Deletions Done')
  
  intTime1 = datetime.now()
  print('Time Taken:\t{}'.format(intTime1 - intTime2))
  intTime2 = intTime1
  
  # Split "All Course Summaries" into separate sheet for each course
  for i_crse, crse in df_prg_courses.iterrows():
    # copy worksheet
    ws_course = wb.copy_worksheet(wb['All Course Summaries'])
    ws_course.title = crse['course_code']
    
    # delete rows of other courses
    while ws_course.cell(row=3, column=1).value != crse['course_code']:
      ws_course.delete_rows(3)
    
    i_row = 3
    while ws_course.cell(row=i_row, column=1).value == crse['course_code']:
      i_row += 1
      
    while ws_course.cell(row=i_row, column=1).value != None:
      ws_course.delete_rows(i_row)

    ws_course.protection.set_password('{}'.format(prg['program_code']))
  
  print('Course Sheets Added')
  
  intTime1 = datetime.now()
  print('Time Taken:\t{}'.format(intTime1 - intTime2))
  intTime2 = intTime1
  
  # Lock sheets
  #for sheet in ['PLOs', 'CLOs', 'Program_course_mapping', 'PLOs_to_CLOs',
  #              'Program Summary', 'All Course Summaries']:
    #ws = wb[sheet]
    #ws.protection.set_password('{}'.format(prg['program_code']))
    
  save_filename = 'v1.6\\PLO_CLO_alignment_{}_28Feb2019.xlsx'.format(prg['program_code'])
  wb.save(directory+save_filename)

  print('File Saved')
  intTime1 = datetime.now()
  print('Time Taken:\t{}'.format(intTime1 - intTime2))
  intTime2 = intTime1
  
  print('{}\tTime Taken:\t{}'.format(prg['program_code'], intTime2 - startTime))
  startTime = intTime2
  print('\n\n')
