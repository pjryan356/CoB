## Creates course assessment files for CAC purposes
# Peter Ryan Nov 2018

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from tabulate import tabulate

import sys
sys.path.append('c:\\Peter\\GitHub\\CoB\\')


def adjust_excel_rows(ws, skip_rows=0, max_rows=False, default_col='A', default_height=15):
  width = ws.column_dimensions[default_col].width
  for i_row, row in enumerate(ws):
    if (i_row + 1 > skip_rows and
      (max_rows == False or max_rows >= (i_row + 1 - skip_rows))):
      multiples_of_font_size = [1]
      for j_col, cell in enumerate(row):
        if cell.value is not None:
          mul = 0
          for v in str(cell.value).split('\n'):
            mul += int(len(v)/width) + 1
          if mul > 0:
            multiples_of_font_size.append(mul)
          
        cell.alignment = Alignment(wrap_text=True)
      
      new_height = max(multiples_of_font_size) * default_height
      ws.row_dimensions[i_row + 1].height = new_height


# open template
directory = 'H:\\Projects\\CoB\\Program Transformation\\CLO mapping\\'
plo_clo_filename = 'PLOs_CLOs_manually_editted.xlsx'

template = 'PLO_CLO_alignment_template_1.xlsx'

# Get Dataframes
df_plo = pd.read_excel(directory+plo_clo_filename, sheet_name='PLOs')
df_clo = pd.read_excel(directory+plo_clo_filename, sheet_name='CLOs', converters={'course_id': str})
df_mapping = pd.read_excel(directory+plo_clo_filename, sheet_name='Mapping', converters={'course_id': str})

# get program list
df_programs = df_mapping[['program_code', 'plan_code', 'career', 'school_code',
                          'school_abbr', 'campus', 'program_name']].drop_duplicates()

# iterate through programs
for i_prg, prg in df_programs.iterrows():
  print(prg)
  # open template
  wb = openpyxl.load_workbook(directory + template)

  # Update PLO sheet
  # Open PLO sheet
  ws = wb["PLOs"]
  
  df_prg_plos = df_plo.loc[(df_plo['program_code'] == prg['program_code']) & (df_plo['plan_code'] == prg['plan_code'])]
  
  j = 2
  for i, r in df_prg_plos.iterrows():
    ws.cell(row=j, column=1).value = prg['program_code']
    ws.cell(row=j, column=2).value = prg['plan_code']
    ws.cell(row=j, column=3).value = prg['program_name']
    ws.cell(row=j, column=4).value = r['career']
    ws.cell(row=j, column=5).value = r['school_code']
    ws.cell(row=j, column=6).value = r['school_abbr']
    ws.cell(row=j, column=7).value = r['plo_nbr']
    ws.cell(row=j, column=8).alignment = Alignment(wrapText=True)
    ws.cell(row=j, column=8).value = r['plo_text']
    ws.cell(row=j, column=9).value = r['status']
    ws.cell(row=j, column=10).value = r['updated']
    j += 1

  # Update mapping sheet
  # Open Mapping sheet
  ws = wb["Program_course_mapping"]

  df_prg_mapping = df_mapping.loc[(df_mapping['program_code'] == prg['program_code']) & (df_mapping['plan_code'] == prg['plan_code'])]
  
  j=2
  for i, r in df_prg_mapping.iterrows():
    df_prg_crse_clos = df_clo.loc[(df_clo['course_id'] == r['course_id'])]
    
    ws.cell(row=j, column=1).value = prg['program_code']
    ws.cell(row=j, column=2).value = prg['plan_code']
    ws.cell(row=j, column=3).value = prg['program_name']
    ws.cell(row=j, column=4).value = r['course_id']
    ws.cell(row=j, column=5).value = r['course_code']
    ws.cell(row=j, column=6).value = r['course_name']
    ws.cell(row=j, column=7).value = r['course_list_name']
    ws.cell(row=j, column=8).value = len(df_prg_crse_clos)
    j += 1

  # Update CLO & PLOS to CLOs sheets
  # Open sheets
  ws1 = wb["CLOs"]
  ws2 = wb["PLOs_to_CLOs"]
  # Get courses
  df_prg_courses = df_prg_mapping[['course_id', 'course_code', 'course_name']].drop_duplicates()
  
  j = 2
  k = 2
  for i_crse, crse in df_prg_courses.iterrows():
    df_prg_crse_clos = df_clo.loc[(df_clo['course_id'] == crse['course_id'])]

    for i, clo in df_prg_crse_clos.iterrows():
      ws1.cell(row=j, column=1).value = crse['course_id']
      ws1.cell(row=j, column=2).value = crse['course_code']
      ws1.cell(row=j, column=3).value = crse['course_name']
      ws1.cell(row=j, column=4).value = clo['school_code']
      ws1.cell(row=j, column=5).value = clo['school_abbr']
      ws1.cell(row=j, column=6).value = clo['clo_nbr']
      ws1.cell(row=j, column=7).alignment = Alignment(wrapText=True)
      ws1.cell(row=j, column=7).value = clo['clo_text']
      ws1.cell(row=j, column=8).value = clo['status']
      ws1.cell(row=j, column=9).value = clo['updated']
      j += 1
      # put all PLO combination in PLOs to CLOs
      for i_plo, plo in df_prg_plos.iterrows():
        ws2.cell(row=k, column=1).value = crse['course_code']
        ws2.cell(row=k, column=2).value = clo['clo_nbr']
        ws2.cell(row=k, column=3).value = plo['plo_nbr']
        ws2.cell(row=k, column=4).alignment = Alignment(wrapText=True)
        ws2.cell(row=k, column=4).value = clo['clo_text']
        ws2.cell(row=k, column=5).alignment = Alignment(wrapText=True)
        ws2.cell(row=k, column=5).value = plo['plo_text']
        k += 1
  
  # Adjust row heights and column widths
  ws = wb["Program Summary"]
  ws.column_dimensions['B'].width = 60
  adjust_excel_rows(ws, skip_rows=3, default_col='B', default_height=15)

  ws = wb["All Course Summaries"]
  ws.row_dimensions[2].height = 200
  for col in ['E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
    ws.column_dimensions[col].width = 16
  for row in range(3, 2000):
    ws.row_dimensions[row].height = 50

  ws = wb["PLOs"]
  ws.column_dimensions['H'].width = 80
  adjust_excel_rows(ws, skip_rows=1, default_col='H', default_height=15)
  
  ws = wb["CLOs"]
  ws.column_dimensions['G'].width = 80
  adjust_excel_rows(ws, skip_rows=1, default_col='G', default_height=15)
  
  ws = wb["PLOs_to_CLOs"]
  ws.column_dimensions['D'].width = 80
  ws.column_dimensions['E'].width = 80
  adjust_excel_rows(ws, skip_rows=1, default_col='D', default_height=16)
  
  
  
  save_filename = '\\Done\\PL0_CLO_alignment_{}_{}.xlsx'.format(prg['school_abbr'], prg['program_code'])
  wb.save(directory+save_filename)