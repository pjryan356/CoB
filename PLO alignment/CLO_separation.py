## Creates course assessment files for CAC purposes
# Peter Ryan Nov 2018

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import sys
import re
sys.path.append('c:\\Peter\\GitHub\\CoB\\')


def get_school_name(school_code):
  if school_code == '610P':
    return 'CPO'
  if school_code == '615H':
    return 'ACCT'
  if school_code == '620H':
    return 'BITL'
  if school_code == '625H':
    return 'EFM'
  if school_code == '630H':
    return 'MGT'
  if school_code == '650T':
    return 'VBE'
  if school_code == '660H':
    return 'GSBL'
  if school_code == 'VN':
    return 'SBM'
  return None


def cleanhtml(raw_html):
  cleanr = re.compile('<.*?>')
  cleantext = re.sub(cleanr, '', raw_html)
  return cleantext

def get_CLOs(text, cid=None):
  try:
    if '<li>' in text or '</li>' in text:
      CLOs = re.split('<li>|</li>|CLO|<p>|</p>|<div>|</div>|', text)
      if len(CLOs) == 1:
        CLOs = re.split('<li>|</li>|<br>|<br />|</p>|<p>|</br>|<div>|</div>|CLO', text)
      
      CLOs_2 = []
      for clo in CLOs:
        clo = clo.replace('<br />', '\n')
        clo = cleanhtml(clo)
        try:
          clo = clo.strip()
          clo = clo.rstrip()
          clo = clo.strip('-')
          clo = clo.strip('•')
          clo = clo.strip('*')
          clo = clo.strip('1')
          clo = clo.strip('2')
          clo = clo.strip('3')
          clo = clo.strip('4')
          clo = clo.strip('5')
          clo = clo.strip('6')
          clo = clo.strip('7')
          clo = clo.strip('8')
          clo = clo.strip('9')
          clo = clo.strip('0')
          clo = clo.strip('.')
          if clo.startswith(":"):
            clo = clo[1:]
          clo = clo.strip(')')
          clo = clo.strip('\uf0a7')
          clo = clo.strip()
        except: pass
      
        if len(clo) > 10:
          CLOs_2.append(clo)

    else:
      NotOKlist = ['<p>', '</p>', '<br>', '<br />', '</br>', '<div>', '</div>','CLO']
      if any(s in text for s in NotOKlist):
        CLOs = re.split('<br>|<br />|</p>|<p>|</br>|CLO|<div>|</div>', text)
      else:
        return [text]
      CLOs_2 = []
      for clo in CLOs:
        clo = clo.replace('<br />', '\n')
        clo = cleanhtml(clo)
        try:
          clo = clo.strip()
          clo = clo.rstrip()
          clo = clo.strip('-')
          clo = clo.strip('•')
          clo = clo.strip('*')
          clo = clo.strip('1')
          clo = clo.strip('2')
          clo = clo.strip('3')
          clo = clo.strip('4')
          clo = clo.strip('5')
          clo = clo.strip('6')
          clo = clo.strip('7')
          clo = clo.strip('8')
          clo = clo.strip('9')
          clo = clo.strip('0')
          clo = clo.strip('.')
          if clo.startswith(":"):
            clo =  clo[1:]
          clo = clo.strip(')')
          clo = clo.strip('\uf0a7')
          clo = clo.strip()
        except Exception as e:
          print(cid)
          print(e)
          pass
        
        if len(clo) > 10:
          CLOs_2.append(clo)

    if len(CLOs_2) == 1:
      return CLOs_2
    
    while ':' in CLOs_2[0][-4:]:
      CLOs_2 = CLOs_2[1:]
      
    while 'Learning Outcomes' in CLOs_2[0]:
      CLOs_2 = CLOs_2[1:]

    if 'Learning outcomes' in CLOs_2[0]:
      CLOs_2 = CLOs_2[1:]
    
    if 'learning outcomes' in CLOs_2[0]:
      CLOs_2 = CLOs_2[1:]
    
    if 'completion of this course' in CLOs_2[0]:
      CLOs_2 = CLOs_2[1:]

    if 'Enabling Knowledge and Skills for Capabilities' in CLOs_2[0]:
      CLOs_2 = CLOs_2[1:]
    
    if 'engage in activities leading to an understanding of\n' == CLOs_2[0]:
      CLOs_2 = CLOs_2[1:]
    
    return CLOs_2
  
  except Exception as e:
    print(cid)
    print(e)
    print(text, '\n')
    return['']

# open template
directory = 'H:\\Projects\\CoB\\Program Transformation\\CLO mapping\\Success\\'
clo_filename = 'CLOs_cob_success.xlsx'
template = 'CLO_template_success.xlsx'
savefile = 'CLOs_cob_success_2_extra.xlsx'


# open template
wb = openpyxl.load_workbook(directory+template)


# fill CLOs worksheet
clo_df = pd.read_excel(open(directory+clo_filename, 'rb'), converters={'Course ID': str})


clo_ws = wb.active

j = 2
for i, r in clo_df.iterrows():
  CLO_list = []
  CLO_list = get_CLOs(r['Learning Outcomes'], r['Course ID'])
  k = 1
  for clo in CLO_list:
    clo_ws.cell(row=j, column=1).value = r['Course ID']
    clo_ws.cell(row=j, column=2).value = r['Course Title']
    clo_ws.cell(row=j, column=3).value = r['School ID']
    clo_ws.cell(row=j, column=4).value = get_school_name(r['School ID'])
    clo_ws.cell(row=j, column=5).value = 'CLO{}'.format(k)
    clo_ws.cell(row=j, column=6).alignment = Alignment(wrapText=True)
    clo_ws.cell(row=j, column=6).value = clo
    clo_ws.cell(row=j, column=7).value = r['Version']
    clo_ws.cell(row=j, column=8).value = r['Status']
    clo_ws.cell(row=j, column=9).value = r['Publish/Unpublish Time']
    j += 1
    k += 1
wb.save(directory+savefile)



