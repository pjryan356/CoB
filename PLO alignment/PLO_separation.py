## Creates course assessment files for CAC purposes
# Peter Ryan Nov 2018

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
from tabulate import tabulate

import sys
import re
sys.path.append('c:\\Peter\\GitHub\\CoB\\')


def get_school_name(school_code):
  if school_code == '610P':
    return 'CBO'
  if school_code == '615H':
    return 'ACCT'
  if school_code == '620H':
    return 'BITL'
  if school_code == '625H':
    return 'EFM'
  if school_code == '630H':
    return 'MGT'
  if school_code == '650T':
    return 'VBE'
  if school_code == '660H':
    return 'GSBL'
  if school_code == 'VN':
    return 'SBM'
  return None


def cleanhtml(raw_html):
  cleanr = re.compile('<.*?>')
  cleantext = re.sub(cleanr, '', raw_html)
  return cleantext

def get_PLOs(text, prog=None):
  try:
    if 'BP141' in prog:
      print(text)
      text = cleanhtml(text)
      print(text)
    PLOs = (text.split('\n'))
    PLOs_2 = []
    for lo in PLOs:
      lo = lo.strip()
      lo = lo.rstrip()
      lo = lo.strip('-')
      lo = lo.strip('•')
      lo = lo.strip('*')
      lo = lo.strip('1')
      lo = lo.strip('2')
      lo = lo.strip('3')
      lo = lo.strip('4')
      lo = lo.strip('5')
      lo = lo.strip('6')
      lo = lo.strip('7')
      lo = lo.strip('8')
      lo = lo.strip('9')
      lo = lo.strip('.')
      if lo.startswith(":"):
        return lo[1:]
      lo = lo.strip(')')
      lo = lo.strip('\uf0a7')
      lo = lo.strip()
      
      if len(lo) > 10:
        PLOs_2.append(lo)
    
    while PLOs_2[0][-1] != ':':
      PLOs_2 = PLOs_2[1:]
    PLOs_2 = PLOs_2[1:]
    return(PLOs_2)
  
  except Exception as e:
    print(e)
    print(prog)
    #print(text)
    return['']

# open template
directory = 'H:\\Projects\\CoB\\Program Transformation\\CLO mapping\\'
clo_filename = 'PLOs_cob_unedited.xlsx'
template = 'PLO_template.xlsx'
savefile = 'PLO_cob.xlsx'


# open template
wb = openpyxl.load_workbook(directory+template)

# fill CLOs worksheet
plo_df = pd.read_excel(open(directory+clo_filename, 'rb'))
plo_ws = wb.active

j = 2
for i, r in plo_df.iterrows():
  if r['Status'] in ['Republished', 'Published']:
    school_code = r['Owning School'].split('(')[1][:-1]
    PLO_list = get_PLOs(r['Statement of Capabilities'], '{} {}'.format(r['Program Code'], r['Plan Code']))
    k = 1
    for plo in PLO_list:
      plo_ws.cell(row=j, column=1).value = r['Program Code']
      plo_ws.cell(row=j, column=2).value = r['Plan Code']
      plo_ws.cell(row=j, column=3).value = r['Program Name']
      plo_ws.cell(row=j, column=4).value = r['Career']
      plo_ws.cell(row=j, column=5).value = school_code
      plo_ws.cell(row=j, column=6).value = get_school_name(school_code)
      plo_ws.cell(row=j, column=7).value = 'PLO{}'.format(k)
      plo_ws.cell(row=j, column=8).alignment = Alignment(wrapText=True)
      plo_ws.cell(row=j, column=8).value = plo
      plo_ws.cell(row=j, column=9).value = r['Status']
      plo_ws.cell(row=j, column=10).value = r['Created Date']
      j += 1
      k += 1
wb.save(directory+savefile)



