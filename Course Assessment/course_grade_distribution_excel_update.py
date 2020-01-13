## Creates course assessment files for CAC purposes
# Peter Ryan Nov 2018

import pandas as pd
import openpyxl
import datetime as dt
import shutil

import sys
sys.path.append('c:\\Peter\\GitHub\\CoB\\')
from general.sams_helper_functions import *
from general.sams_queries import *


# Create connections
# create sams engine this is the connection to the oracle database
password_str = input("SAMS Password: ") ## Input password
sams_engine = return_sams_engine(password_str=password_str)

# Get current semester information
current_year = 2019
#current_year = int(input("Current year: "))

current_semester = 2
#current_semester = int(input("Current semester (1 or 2 or 3): "))

#location = input("Location (MELB, SBM, SIM): ")
#level = input("Level (VE, HE): ")
#term_code_final = input("Final term code: ")
location = 'MELB'
level = 'VE'
term_code_final = '1945'

equivalent_semesters = False
#equivalent_semesters = bool(input("Use equivalent semester (True or False): "))

st_year = 2015
#st_year = int(input("Earliest Year: "))

sheet_pw = 'ADG'
#sheet_pw = input("Select Password: ")


def get_school_name(school_code):
  if school_code == '610P':
    return 'CBO'
  if school_code == '615H':
    return 'ACCT'
  if school_code == '620H':
    return 'BITL'
  if school_code == '625H':
    return 'EFM'
  if school_code == '630H':
    return 'MGT'
  if school_code == '650T':
    return 'VBE'
  if school_code == '660H':
    return 'GSBL'
  if school_code == 'VN':
    return 'Not CoB'
  return None

def list_to_text(obList):
  # converts a list of object into a string list for sql IN statement
  txt = "("
  for ob in obList:
    txt += "'{}',".format(ob)
  txt = txt[:-1] + ")"
  return txt

def copy_rename(old_file_name, new_file_name):
  import os
  
  src_dir = os.curdir
  dst_dir = os.path.join(os.curdir, "subfolder")
  src_file = os.path.join(src_dir, old_file_name)
  shutil.copy(src_file, dst_dir)
  
  dst_file = os.path.join(dst_dir, old_file_name)
  new_dst_file_name = os.path.join(dst_dir, new_file_name)
  os.rename(dst_file, new_dst_file_name)

def get_all_bus_courses(term_year, term_code):
  qry = '''
SELECT DISTINCT
    cls.course_code,
    cls.course_name,
    cls.school_code,
    cls.term_code,
    term.term_name

FROM (
    SELECT DISTINCT
      t2.STRM AS term_code,
      t2.SUBJECT || t2.CATALOG_NBR AS course_code,
      t2.DESCR AS course_name,
      t2.acad_org AS school_code

    FROM PS_CLASS_TBL t2
    WHERE t2.acad_group = 'BUS' AND t2.strm={1} AND enrl_tot > 0
    ) cls
  INNER JOIN (
    SELECT DISTINCT
      t3.STRM AS term_code,
      t3.DESCRSHORT AS term_name
    FROM PS_TERM_TBL t3
    WHERE acad_year = {0}
    ) term ON (cls.term_code = term.term_code)
'''.format(term_year, term_code)
  print(qry)
  return qry
  
def get_course_grade_distribution(course_code, term_code='1850', st_year=2015, term_codes=None):
  st_term = '{}00'.format(str(int(st_year)-2000))
  
  term_txt = list_to_text(term_codes)

  qry = '''
SELECT
	cls.course_code,
  enrl.term_code,
  term.term_name,
  sum(enrl.nn) AS nn,
  sum(enrl.pa) AS pa,
  sum(enrl.cr) AS cr,
  sum(enrl.di) AS di,
  sum(enrl.hd) AS hd

FROM (
  SELECT
    t1.CLASS_NBR,
    t1.STRM AS term_code,
    sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 0 AND t1.CRSE_GRADE_OFF IN ('NN', 'DNS', 'NH', 'SP') THEN 1 ELSE 0 END) AS nn,
    sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 1 THEN 1 ELSE 0 END) AS pa,
    sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 2 THEN 1 ELSE 0 END) AS cr,
    sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 3 THEN 1 ELSE 0 END) AS di,
    sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 4 THEN 1 ELSE 0 END) AS hd

  FROM	PS_STDNT_ENRL t1
  WHERE
      t1.strm >= '{0}' AND t1.strm < '{1}' AND substr(strm,3,2) IN {4}
      AND t1.INCLUDE_IN_GPA = 'Y'
      AND t1.STDNT_ENRL_STATUS = 'E'
      AND t1.ENRL_STATUS_REASON='ENRL'
      AND t1.GRADING_BASIS_ENRL<>'NON'

  GROUP BY t1.CLASS_NBR, t1.STRM
  ) enrl

INNER JOIN (
  SELECT DISTINCT
    t2.CLASS_NBR,
    t2.STRM AS term_code,
    t2.SUBJECT || t2.CATALOG_NBR AS course_code
  FROM PS_CLASS_TBL t2
  WHERE t2.acad_group = 'BUS' AND t2.SUBJECT || t2.CATALOG_NBR = '{2}'
  ) cls ON (cls.CLASS_NBR = enrl.CLASS_NBR AND cls.term_code = enrl.term_code)

INNER JOIN (
  SELECT DISTINCT
    t3.STRM AS term_code,
    t3.descrshort AS term_name,
    t3.term_category,
    t3.acad_year AS term_year
  FROM PS_TERM_TBL t3
  WHERE t3.acad_year >= {3}
  ) term ON (enrl.term_code = term.term_code)
GROUP BY cls.course_code, enrl.term_code, term.term_name
ORDER BY enrl.term_code DESC
  '''.format(st_term, term_code, course_code, st_year, term_txt)
  return(qry)

def get_term_category(location, semester=None):
  if semester != None:
    if location == 'MELB':
      return [current_semester]
    if location == 'SBM':
      return [current_semester+4]
    if location == 'SIM':
      if semester == 1:
        return[5]
      if semester == 2:
        return [7]
    else:
      print('Selection not available')
      return None

  if semester == None:
    if location == 'MELB':
      return [1, 2]
    if location == 'SBM':
      return [5, 6, 7]
    if location == 'SIM':
      return [5, 7]
    else:
      print('Selection not available')
      return None

def get_term_code(location, semester=None, level='VE'):
  if semester != None:
    if location == 'MELB':
      
      if semester == 1:
        if level == 'VE':
          return ['05']
        else:
          return ['10']
    
      if semester == 2:
        if level == 'VE':
          return ['45']
        else:
          return ['50']
    
    if location == 'SBM':
      return ['9{}'.format(semester)]

    if location == 'SIM':
      if semester == 1:
        return ['20']
      if semester == 2:
        return ['60']
    else:
      print('Selection not available')
      return None

  if semester == None:
    if location == 'MELB':
      if level == 'VE':
        return ['05', '45']
      else:
        return ['10', '50']
    if location == 'SBM':
      return ['91', '92', '93']
    if location == 'SIM':
      return ['20', '60']
    else:
      print('Selection not available')
      return None
    

term_cat = get_term_category(location, current_semester)
term_code = get_term_code(location, current_semester)


start = dt.datetime.now()
# Get all courses
df_courses = pd.read_sql(sql=get_all_bus_courses(current_year, term_code_final), con=sams_engine)
print (len(df_courses))

# Iterate through courses
for i_course, r_course in df_courses.iterrows():
  #if i_course > 10:
  #  break
  # get course data from sams
  if equivalent_semesters:
    sams_qry = get_course_grade_distribution(course_code=r_course['course_code'],
                                             term_code=r_course['term_code'],
                                             st_year=st_year,
                                             get_term_code=get_term_code(location, current_semester, level))
  else:
    sams_qry = get_course_grade_distribution(course_code=r_course['course_code'],
                                             term_code=r_course['term_code'],
                                             st_year=st_year,
                                             term_codes=get_term_code(location, level=level))
  try:
    df = pd.read_sql(sql=sams_qry, con=sams_engine)
    #print(tabulate(df, headers='keys'))
  except:
    print(sams_qry)
  
  # open template
  directory = 'H:\\Projects\\CoB\\Course_Assessment_Moderation\\2019S2\\'
  template = 'grade_distribution_template.xlsx'
  wb = openpyxl.load_workbook(directory+template)
  sheet = wb['data']

  # Input data into sheet
  header = '{}: {}'.format(r_course['course_code'],
                           r_course['course_name'])
  sheet.cell(row=1, column=3).value = header
  sheet.cell(row=6, column=2).value = r_course['term_name']
  
  for i, r in df.iterrows():
    if i < 3:
      sheet.cell(row=5 - i, column=2).value = r['term_name']
      sheet.cell(row=5 - i, column=3).value = r['nn']
      sheet.cell(row=5 - i, column=4).value = r['pa']
      sheet.cell(row=5 - i, column=5).value = r['cr']
      sheet.cell(row=5 - i, column=6).value = r['di']
      sheet.cell(row=5 - i, column=7).value = r['hd']

  school_name = get_school_name(r_course['school_code'])
  
  # protect sheet
  sheet.protection.set_password('{}'.format(sheet_pw))
  # Save sheet
  filename = '{1}\\{0}_{2}_{3}_grade_distribution.xlsx'.format(school_name, location, r_course['course_code'], r_course['term_name'])
  print(directory+filename)
  wb.save(directory+filename)
  del df
  if i_course%10 == 0:
    print('{} courses in {} seconds'.format(i_course, (dt.datetime.now() - start).seconds))

print('{} courses in {} seconds'.format(i_course, (dt.datetime.now() - start).seconds))

