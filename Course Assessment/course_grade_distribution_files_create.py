# Creates course assessment files for CAC purposes
# Updated Peter Ryan Jan 2020

import pandas as pd
import openpyxl
import datetime as dt
import shutil
from tabulate import tabulate

import sys
sys.path.append('c:\\Peter\\GitHub\\CoB\\')
from general.sams_helper_functions import *
from general.sams_queries import *
from general.db_helper_functions import (
  objectlist_to_text,
  get_school_name,
  get_campus_list_string)


# Input parameters
current_year = 2019
current_semester = 2

# Locations MELB, SBM, SIM, CSI (SUIBE)
location = 'SUIBE'
level = 'HE'

# Term Codes
#   MELB HE (yy10, yy50)
#   MELB VE (yy05, yy45)
#   SBM     (yy91, yy92, yy93)
#   SIM     (yy20, yy60)
#   CSI     (yy08, yy48)
term_code_final = '1948'

equivalent_semesters = False

st_year = 2016
sheet_pw = 'ADG'

# Create database connections
# create sams engine this is the connection to the oracle database
password_str = input("SAMS Password: ")   # Input password
sams_engine = return_sams_engine(password_str=password_str)


def get_all_bus_courses(term_year, term_code, campus=None):
  qry = '''
  SELECT DISTINCT
      cls.course_code,
      cls.course_name,
      cls.school_code,
      cls.term_code,
      term.term_name
  
  FROM (
      SELECT DISTINCT
        t2.STRM AS term_code,
        t2.SUBJECT || t2.CATALOG_NBR AS course_code,
        t2.DESCR AS course_name,
        t2.acad_org AS school_code
  
      FROM PS_CLASS_TBL t2
      WHERE t2.acad_group = 'BUS' AND t2.strm={0} AND enrl_tot > 0
  '''.format(term_code)
  if campus != None:
    qry += ' AND campus IN {} '.format(campus)
  qry += '''
      ) cls
    INNER JOIN (
      SELECT DISTINCT
        t3.STRM AS term_code,
        t3.DESCRSHORT AS term_name
      FROM PS_TERM_TBL t3
      WHERE acad_year = {0}
      ) term ON (cls.term_code = term.term_code)
  '''.format(term_year)
  return qry
  

def get_course_grade_distribution(course_code, term_code='1850', start_year=2015, term_codes_ends=None):
  start_term = '{}00'.format(str(int(start_year)-2000))

  term_end_txt = objectlist_to_text(term_codes_ends)

  qry = '''
  SELECT
    cls.course_code,
    enrl.term_code,
    term.term_name,
    sum(enrl.nn) AS nn,
    sum(enrl.pa) AS pa,
    sum(enrl.cr) AS cr,
    sum(enrl.di) AS di,
    sum(enrl.hd) AS hd
  
  FROM (
    SELECT
      t1.CLASS_NBR,
      t1.STRM AS term_code,
      sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 0 AND t1.CRSE_GRADE_OFF IN ('NN', 'DNS', 'NH', 'SP') THEN 1 ELSE 0 END) AS nn,
      sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 1 THEN 1 ELSE 0 END) AS pa,
      sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 2 THEN 1 ELSE 0 END) AS cr,
      sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 3 THEN 1 ELSE 0 END) AS di,
      sum(CASE WHEN t1.GRD_PTS_PER_UNIT = 4 THEN 1 ELSE 0 END) AS hd
  
    FROM	PS_STDNT_ENRL t1
    WHERE
        t1.strm >= '{0}' AND t1.strm < '{1}' AND substr(strm,3,2) IN {4}
        AND t1.INCLUDE_IN_GPA = 'Y'
        AND t1.STDNT_ENRL_STATUS = 'E'
        AND t1.ENRL_STATUS_REASON='ENRL'
        AND t1.GRADING_BASIS_ENRL<>'NON'
  
    GROUP BY t1.CLASS_NBR, t1.STRM
    ) enrl
  
  INNER JOIN (
    SELECT DISTINCT
      t2.CLASS_NBR,
      t2.STRM AS term_code,
      t2.SUBJECT || t2.CATALOG_NBR AS course_code
    FROM PS_CLASS_TBL t2
    WHERE t2.acad_group = 'BUS' AND t2.SUBJECT || t2.CATALOG_NBR = '{2}'
    ) cls ON (cls.CLASS_NBR = enrl.CLASS_NBR AND cls.term_code = enrl.term_code)
  
  INNER JOIN (
    SELECT DISTINCT
      t3.STRM AS term_code,
      t3.descrshort AS term_name,
      t3.term_category,
      t3.acad_year AS term_year
    FROM PS_TERM_TBL t3
    WHERE t3.acad_year >= {3}
    ) term ON (enrl.term_code = term.term_code)
  GROUP BY cls.course_code, enrl.term_code, term.term_name
  ORDER BY enrl.term_code DESC
  '''.format(start_term, term_code, course_code, start_year, term_end_txt)
  return qry


start = dt.datetime.now()

# Get all courses
df_courses = pd.read_sql(sql=get_all_bus_courses(current_year,
                                                 term_code_final,
                                                 get_campus_list_string(location)
                                                 ),
                         con=sams_engine)
print(tabulate(df_courses, headers='keys'))

# Iterate through courses
for i_course, r_course in df_courses.iterrows():
  # While testing new semester turn this on
  # if i_course > 4:
  #   break
  
  # Get course data from sams
  if equivalent_semesters:
    sams_qry = get_course_grade_distribution(course_code=r_course['course_code'],
                                             term_code=r_course['term_code'],
                                             start_year=st_year,
                                             term_codes_ends=get_term_code_ends(location,
                                                                                current_semester,
                                                                                level))
  else:
    sams_qry = get_course_grade_distribution(course_code=r_course['course_code'],
                                             term_code=r_course['term_code'],
                                             start_year=st_year,
                                             term_codes_ends=get_term_code_ends(location,
                                                                                None,
                                                                                level))
  try:
    df = pd.read_sql(sql=sams_qry, con=sams_engine)
  except:
    print(sams_qry)
  
  # open template
  directory = 'H:\\Projects\\CoB\\Course_Assessment_Moderation\\{}S{}\\'.format(current_year, current_semester)
  template = 'grade_distribution_template.xlsx'
  wb = openpyxl.load_workbook(directory+template)
  sheet = wb['data']

  # Input data into sheet
  header = '{}: {}'.format(r_course['course_code'],
                           r_course['course_name'])
  sheet.cell(row=1, column=3).value = header
  sheet.cell(row=6, column=2).value = r_course['term_name']
  
  for i, r in df.iterrows():
    if i < 3:
      sheet.cell(row=5 - i, column=2).value = r['term_name']
      sheet.cell(row=5 - i, column=3).value = r['nn']
      sheet.cell(row=5 - i, column=4).value = r['pa']
      sheet.cell(row=5 - i, column=5).value = r['cr']
      sheet.cell(row=5 - i, column=6).value = r['di']
      sheet.cell(row=5 - i, column=7).value = r['hd']

  school_name = get_school_name(r_course['school_code'])
  
  # protect sheet
  sheet.protection.set_password('{}'.format(sheet_pw))
  # Save sheet
  filename = '{1}\\{0} {2} {3} grade distribution.xlsx'.format(school_name,
                                                               location,
                                                               r_course['course_code'],
                                                               r_course['term_name'])
  print(directory+filename)
  wb.save(directory+filename)
  del df
  if i_course % 10 == 0:
    print('{} courses in {} seconds'.format(i_course, (dt.datetime.now() - start).seconds))
