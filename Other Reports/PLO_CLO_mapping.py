## Creates course assessment files for CAC purposes
# Peter Ryan Nov 2018

import pandas as pd
import openpyxl
from openpyxl.styles import Alignment
import datetime as dt
import shutil
from tabulate import tabulate

import sys
sys.path.append('c:\\Peter\\GitHub\\CoB\\')

program_code = 'BP251'

def get_school_name(school_code):
  if school_code == '610P':
    return 'CPO'
  if school_code == '615H':
    return 'ACCT'
  if school_code == '620H':
    return 'BITL'
  if school_code == '625H':
    return 'EFM'
  if school_code == '630H':
    return 'MGT'
  if school_code == '650T':
    return 'VBE'
  if school_code == '660H':
    return 'GSBL'
  if school_code == 'VN':
    return 'SBM'
  return None



def copy_rename(old_file_name, new_file_name):
  import os
  
  src_dir = os.curdir
  dst_dir = os.path.join(os.curdir, "subfolder")
  src_file = os.path.join(src_dir, old_file_name)
  shutil.copy(src_file, dst_dir)
  
  dst_file = os.path.join(dst_dir, old_file_name)
  new_dst_file_name = os.path.join(dst_dir, new_file_name)
  os.rename(dst_file, new_dst_file_name)

def get_CLOs(text):
  CLOs = text.split('<li>')
  CLO_list = []
  for clo in CLOs[1:]:
    clo = clo.replace('<br />', '\n')
    clo = clo.split('</li>')[0]
    CLO_list.append(clo)
  return CLO_list

# open template
directory = 'H:\\Projects\\CoB\\Program Transformation\\2019\\'
clo_filename = 'CLOs_{}.xlsx'.format(program_code)
plo_filename = 'PLOs_{}.xlsx'.format(program_code)
mapping_filename = 'Active_programs_core_course_list_2018.xlsx'
template = 'PLO_CLO_alignment_template.xlsx'
savefile = 'PLO_CLO_alignment_template_{}.xlsx'.format(program_code)


# open template
wb = openpyxl.load_workbook(directory+template)


# fill CLOs worksheet
clo_df = pd.read_excel(open(directory+filename, 'rb'))
clo_ws = wb["CLOs"]

j = 2
for i, r in clo_df.iterrows():
  CLO_list = get_CLOs(r[6])
  k = 1
  for clo in CLO_list:
    clo_ws.cell(row=j, column=1).value = r[0]
    clo_ws.cell(row=j, column=2).value = r[1]
    clo_ws.cell(row=j, column=3).value = r[3]
    clo_ws.cell(row=j, column=4).value = get_school_name(r[3])
    clo_ws.cell(row=j, column=5).value = 'CLO{}'.format(k)

    clo_ws.cell(row=j, column=6).alignment = Alignment(wrapText=True)
    clo_ws.cell(row=j, column=6).value = clo
    j += 1
    k += 1
wb.save(directory+savefile)

# fill PLOs worksheet
plo_df = pd.read_excel(open(directory+filename, 'rb'))
plo_ws = wb["PLOs"]

j = 2
for i, r in plo_df.iterrows():
  PLO_list = get_PLOs(r[6])
  k = 1
  for plo in PLO_list:
    plo_ws.cell(row=j, column=1).value = r[0]
    plo_ws.cell(row=j, column=2).value = r[1]
    plo_ws.cell(row=j, column=3).value = r[3]
    plo_ws.cell(row=j, column=4).value = get_school_name(r[3])
    plo_ws.cell(row=j, column=5).value = 'CLO{}'.format(k)

    plo_ws.cell(row=j, column=6).alignment = Alignment(wrapText=True)
    plo_ws.cell(row=j, column=6).value = plo
    j += 1
    k += 1

wb.save(directory+savefile)

# Fill Course Mapping worksheet
mapping_df = pd.read_excel(open(directory+mapping_filename, 'rb'))
clo_ws = wb["Mapping"]



wb.save(directory+savefile)


