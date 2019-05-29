import base64

import pandas as pd
import openpyxl
import datetime as dt
import shutil

import sys
sys.path.append('c:\\Peter\\GitHub\\CoB\\')

import general.RMIT_colours as rc

from general.db_helper_functions import (
  connect_to_postgres_db,
  db_extract_query_to_dataframe
)


'''------------------------------------- Get Inputs  --------------------------------'''
# Set parameter values with input prompts or go with preset values (input prompt)
set_values = input("Do you want to manually input the semester/term variables [Y/N]: ")
if set_values == 'Y':
  year = input("Year: ")
  semester = input("Semester: ")
else:
  year = 2019
  semester = 1

'''--------------------------------- Connect to Database  ----------------------------'''
# create postgres engine this is the connection to the postgres database
postgres_pw = input("Postgres Password: ")
postgres_user = 'pjryan'
postgres_host = 'localhost'
postgres_dbname = 'postgres'

con_string = "host='{0}' " \
             "dbname='{1}' " \
             "user='{2}' " \
             "password='{3}' " \
             "".format(postgres_host, postgres_dbname, postgres_user, postgres_pw)

postgres_con, postgres_cur = connect_to_postgres_db(con_string)

'''------------------------Get Images---------------------'''
# header image
image_filename = 'C:\\Peter\\CoB\\logos\\L&T_Transparent_200.png'  # replace with your own image
logo = base64.b64encode(open(image_filename, 'rb').read())

# ces scale image (3 explanation)
image_filename = 'C:\\Peter\\CoB\\logos\\CES_scale.png'  # replace with your own image
ces_scale_image = base64.b64encode(open(image_filename, 'rb').read())


'''------------------------------ Helper functions -----------------------------------'''
def get_last_semester(year, semester):
  # returns the previous year and semester based on a 2 semester year
  if semester != 1:
    semester -= 1
  else:
    semester = 2
    year -= 1
  return year, semester


def list_to_text(obList):
  # converts a list of object into a string list for sql IN statement
  txt = "("
  for ob in obList:
    txt += "'{}',".format(ob)
  txt = txt[:-1] + ")"
  return txt


'''----------------------------- create data extraction functions -------------------------------------'''

def qry_course_teacher_data(year, semester, tbl='vw001_course_teacher', schema='sim_ces'):
  # Returns a dataframe of the sim ces data for year and semester
  qry = " SELECT *  \n" \
        " FROM {0}.{1} \n" \
        " WHERE year = {2} AND semester = {3} \n" \
        "".format(schema, tbl, year, semester)
  return qry

def get_course_teacher_data(year, semester, cur, tbl='vw001_course_teacher', schema='sim_ces'):
  # Returns a dataframe of the SIM ces in year, semester from db (cur)
  qry = qry_course_teacher_data(year, semester, tbl, schema)
  return db_extract_query_to_dataframe(qry, cur, print_messages=False)


def get_course_comments(year, semester, cur, tbl='vw001_course_teacher_comments', schema='sim_ces'):
  # Returns a dataframe with SIM CES comments for courses in year and semester
  qry = " SELECT *  \n" \
        " FROM {0}.{1} \n" \
        " WHERE year = {2} AND semester = {3} \n" \
        "".format(schema, tbl, year, semester)
  return db_extract_query_to_dataframe(qry, cur, print_messages=False)

def filter_teacher_data(df1, school, teacher):
  # filters dataframe to given course_code
  try:
    return df1.loc[(df1['teaching_staff'] == teacher) & (df1['school'] == school)]
  except:
    pass
  return None

def filter_course_data(df1, school, course_code):
  # filters dataframe to given course_code
  try:
    return df1.loc[(df1['course_code'] == course_code) & (df1['school'] == school)]
  except:
    pass
  return None

'''-------------------------------------------- Create Dataframes -------------------------------------'''
df_ces = get_course_teacher_data(year, semester,
                                 cur=postgres_cur)
print(len(df_ces))


df_ces_comments = get_course_comments(year, semester,
                                      cur=postgres_cur)

df_teachers = df_ces[['year', 'semester', 'school', 'teaching_staff']].drop_duplicates()
print(len(df_teachers))


i_teacher_cnt = 0
for i_teacher, r_teacher in df_teachers.iterrows():
  i_teacher_cnt += 1
  print(i_teacher_cnt, i_teacher, r_teacher['teaching_staff'])
  df_ces_teacher_data = filter_teacher_data(df_ces, r_teacher['school'], r_teacher['teaching_staff'])
  df_ces_teacher_comm = filter_teacher_data(df_ces_comments, r_teacher['school'], r_teacher['teaching_staff'])
    

  # open template
  data_directory = 'C:\\Peter\\GitHub\\CoB\\SIM\\'
  save_directory = 'H:\\Projects\\CoB\\CES\\SIM\\'
  template = 'SIM_teacher_template.xlsx'
  wb = openpyxl.load_workbook(data_directory + template)
  data_sheet = wb['Data']
  comm_sheet = wb['Comments']

  # Input headers
  header = '{2} ({0} Semester {1})'.format(r_teacher['year'], r_teacher['semester'],
                                           r_teacher['teaching_staff'])
  data_sheet.cell(row=1, column=2).value = header
  comm_sheet.cell(row=1, column=2).value = header
  
  i_data = 0
  for i, r in df_ces_teacher_data.iterrows():
    i_data += 1
    data_sheet.cell(row=2 + i_data, column=1).value = r['school']
    data_sheet.cell(row=2 + i_data, column=2).value = r['course_code']
    data_sheet.cell(row=2 + i_data, column=3).value = r['course_name']
    data_sheet.cell(row=2 + i_data, column=4).value = r['section_code']
    data_sheet.cell(row=2 + i_data, column=5).value = r['population']
    data_sheet.cell(row=2 + i_data, column=6).value = r['responses']
    data_sheet.cell(row=2 + i_data, column=7).value = round(100.0*r['responses']/r['population'], 1)
    data_sheet.cell(row=2 + i_data, column=8).value = r['subject_content']
    data_sheet.cell(row=2 + i_data, column=9).value = r['lecturer_effectiveness']
    data_sheet.cell(row=2 + i_data, column=10).value = r['course_satisfaction']
    
  i_comm = 0
  for i, r in df_ces_teacher_comm.iterrows():
    i_comm += 1
    comm_sheet.cell(row=2 + i_comm, column=1).value = r['school']
    comm_sheet.cell(row=2 + i_comm, column=2).value = r['course_code']
    comm_sheet.cell(row=2 + i_comm, column=3).value = r['course_name']
    comm_sheet.cell(row=2 + i_comm, column=4).value = r['section_code']
    comm_sheet.cell(row=2 + i_comm, column=5).value = r['comment_type']
    comm_sheet.cell(row=2 + i_comm, column=6).value = r['comment_text']
  
  # protect sheet
  #data_sheet.protection.set_password('{}'.format('data'))
  #comm_sheet.protection.set_password('{}'.format('comments'))
  
  # Save sheet
  filename = '{0}\\SIM_{0}S{1}_{2}_{3}_teacher_evaluation_data.xlsx'.format(year, semester, r['school'], r['teaching_staff'])
  filename = filename.replace('/', '-')
  try:  wb.save(save_directory + filename)
  
  except Exception as e:
    print(save_directory + filename)
    print(e)

df_courses = df_ces[['year', 'semester', 'school', 'course_code']].drop_duplicates()
print(len(df_courses))

i_course_cnt = 0
for i_course, r_course in df_courses.iterrows():
  i_course_cnt += 1
  print(i_course_cnt, i_course, r_course['course_code'])
  df_ces_teacher_data = filter_course_data(df_ces, r_course['school'], r_course['course_code'])
  df_ces_teacher_comm = filter_course_data(df_ces_comments, r_course['school'], r_course['course_code'])
  
  # open template
  data_directory = 'C:\\Peter\\GitHub\\CoB\\SIM\\'
  save_directory = 'H:\\Projects\\CoB\\CES\\SIM\\'
  template = 'SIM_course_template.xlsx'
  wb = openpyxl.load_workbook(data_directory + template)
  data_sheet = wb['Data']
  comm_sheet = wb['Comments']
  
  # Input headers
  header = '{2} ({0} Semester {1})'.format(r_course['year'], r_course['semester'],
                                           r_course['course_code'])
  data_sheet.cell(row=1, column=2).value = header
  comm_sheet.cell(row=1, column=2).value = header
  
  i_data = 0
  for i, r in df_ces_teacher_data.iterrows():
    i_data += 1
    data_sheet.cell(row=2 + i_data, column=1).value = r['school']
    data_sheet.cell(row=2 + i_data, column=2).value = r['course_code']
    data_sheet.cell(row=2 + i_data, column=3).value = r['course_name']
    data_sheet.cell(row=2 + i_data, column=4).value = r['teaching_staff']
    data_sheet.cell(row=2 + i_data, column=5).value = r['section_code']
    data_sheet.cell(row=2 + i_data, column=6).value = r['population']
    data_sheet.cell(row=2 + i_data, column=7).value = r['responses']
    data_sheet.cell(row=2 + i_data, column=8).value = round(100.0 * r['responses'] / r['population'], 1)
    data_sheet.cell(row=2 + i_data, column=9).value = r['subject_content']
    data_sheet.cell(row=2 + i_data, column=10).value = r['lecturer_effectiveness']
    data_sheet.cell(row=2 + i_data, column=11).value = r['course_satisfaction']
  
  i_comm = 0
  for i, r in df_ces_teacher_comm.iterrows():
    i_comm += 1
    comm_sheet.cell(row=2 + i_comm, column=1).value = r['school']
    comm_sheet.cell(row=2 + i_comm, column=2).value = r['course_code']
    comm_sheet.cell(row=2 + i_comm, column=3).value = r['course_name']
    comm_sheet.cell(row=2 + i_comm, column=4).value = r['teaching_staff']
    comm_sheet.cell(row=2 + i_comm, column=5).value = r['section_code']
    comm_sheet.cell(row=2 + i_comm, column=6).value = r['comment_type']
    comm_sheet.cell(row=2 + i_comm, column=7).value = r['comment_text']
  
  # protect sheet
  # data_sheet.protection.set_password('{}'.format('data'))
  # comm_sheet.protection.set_password('{}'.format('comments'))
  
  # Save sheet
  filename = '{0}\\Course_Files\\SIM_{0}S{1}_{2}_{3}_teacher_evaluation_data.xlsx'.format(year, semester, r_course['school'],
                                                                            r_course['course_code'])
  filename = filename.replace('/', '-')
  try:
    wb.save(save_directory + filename)
  
  except Exception as e:
    print(save_directory + filename)
    print(e)
    
'''----------------------------- create dash functions -------------------------------------'''


def make_course_pack(course_code_ces):
  # Main function that creates the Data pack for given course_code
  ## Note the first page header is not included as it forms part of the selection box
  
  # filters data frames to selected course
  df1_ce = get_course_data(df_ce, course_code_ces)
  df1_ces = get_course_data(df_ce_ces, course_code_ces)
  df1_comments = get_course_data(df_ce_comments, course_code_ces)
  df1_themes = get_course_data(df_ce_comment_themes, course_code_ces)
  df1_prg_ces = get_course_data(df_ce_prg_ces, course_code_ces)

 







